"""
Лоадер планов из СВОД ...xlsx, лист «II-Q» (матрица аптеки × проекты).

Структура:
  стр.2 — имена проектов в начале каждого блока (S=BAYER, GF=KRKA, ...)
  стр.3 — заголовки: базовые (B=ИНН, F=юр.название, J=менеджер, K=регион, L=район,
          N=Датфо, O=Plan(IIQ), P=Fact(IIQ)) + повторяющиеся блоки по проекту:
          'проект','ПЛАНIIQ2026','ПЛАНАПРЕЛЬ','ФАКТАПРЕЛЬ',... ['условия']
  стр.4 — строка ИТОГО (ИНН 7777), пропускаем
  стр.5+ — по одной аптеке

Парсим блоки ДИНАМИЧЕСКИ: колонка со значением 'проект' в стр.3 = старт блока,
имя проекта берём из стр.2. Колонки до следующего 'проект' — метрики блока.

Возвращает: dict inn(str) -> {
    meta: {name, network, n_apt, manager, region, district},
    total: {plan_iiq, fact_iiq},
    projects: { '<II-Q имя>': {plan_iiq, plan_apr, plan_may, plan_jun, condition} }
}
"""
import os
import re

import openpyxl
from openpyxl.utils import column_index_from_string

DEFAULT_PATH = os.path.join(os.path.dirname(__file__), 'data', 'plans.xlsx')
SHEET = 'II-Q'
SVOD_SHEET = 'svod'

# Метки внутри блока проекта (стр.3) -> наш ключ
LABEL_MAP = {
    'ПЛАНIIQ2026': 'plan_iiq',
    'ПЛАНАПРЕЛЬ': 'plan_apr',
    'ФАКТАПРЕЛЬ': 'fact_apr',
    'ПЛАНМАЙ': 'plan_may',
    'ФАКТМАЙ': 'fact_may',
    'ПЛАНИЮНЬ': 'plan_jun',
    'ФАКТИЮНЬ': 'fact_jun',
    'ВПIIQ(сум)': 'fact_iiq',   # выполнение плана за квартал в сумме = факт квартала
    'ВПIIQ(%)': 'vp_pct',
    'ДОПЛАНА': 'do_plana',
    'условия': 'condition',
}


def _num(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(' ', '').replace(',', '.'))
    except ValueError:
        return 0.0


def _norm(s):
    """Нормализует заголовок для сравнения: убирает пробелы, дефисы, переносы.
    Дефисы убираем, т.к. в разных выгрузках метки пишут и слитно ('ПЛАНIIQ2026'),
    и через дефис ('ПЛАН -II-Q-2026') — после нормализации они совпадают."""
    return (str(s) if s is not None else '').replace(' ', '').replace('-', '').replace('\n', '').strip()


def _norm_match(s):
    """Как _norm, но ещё и в нижний регистр — для матча базовых заголовков (RU+Latin)."""
    return _norm(s).lower()


# Базовые (мета) колонки матрицы определяем по тексту заголовка в стр.3,
# а не по фиксированным позициям — разные выгрузки сдвигают колонки.
# (наш ключ, предикат на нормализованном заголовке). Первое совпадение слева побеждает.
BASE_MATCHERS = [
    ('inn',      lambda h: h == 'инн'),
    ('legal',    lambda h: 'юридическ' in h),
    ('pharmacy', lambda h: h in ('аптеки', 'аптека')),
    ('network',  lambda h: 'сеть' in h),
    ('n_apt',    lambda h: h.startswith('кол') or 'колво' in h),
    ('manager',  lambda h: h in ('менежер', 'менеджер')),
    ('region',   lambda h: h == 'регион'),
    ('district', lambda h: h == 'район'),
    ('plan_iiq', lambda h: h.startswith('plan')),
    ('fact_iiq', lambda h: h.startswith('fact')),
]


def _detect_base_cols(row3_norm, limit):
    """Находит индексы базовых колонок в стр.3 (сканируем только до первого блока проекта)."""
    cols = {}
    for j in range(min(limit, len(row3_norm))):
        h = row3_norm[j]
        if not h:
            continue
        for key, pred in BASE_MATCHERS:
            if key not in cols and pred(h):
                cols[key] = j
                break
    return cols


def _detect_blocks(ws):
    """Находит блоки проектов: [(project_name_iiq, {our_key: col_idx})]."""
    row2 = [ws.cell(row=2, column=j + 1).value for j in range(ws.max_column)]
    row3 = [ws.cell(row=3, column=j + 1).value for j in range(ws.max_column)]

    # Индексы колонок, где стр.3 == 'проект'
    starts = [j for j, v in enumerate(row3) if _norm(v) == _norm('проект')]
    blocks = []
    for bi, start in enumerate(starts):
        end = starts[bi + 1] if bi + 1 < len(starts) else len(row3)
        name = row2[start] or f'BLOCK_{start}'
        cols = {}
        for j in range(start + 1, end):
            label = _norm(row3[j])
            for raw, key in LABEL_MAP.items():
                if label == _norm(raw):
                    cols[key] = j
        blocks.append((str(name).strip(), cols))
    return blocks


def load_plans(path=DEFAULT_PATH):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET]
    blocks = _detect_blocks(ws)

    # Базовые (мета) колонки определяем по заголовкам стр.3 (позиции плавают между выгрузками).
    row3_vals = [ws.cell(row=3, column=j + 1).value for j in range(ws.max_column)]
    starts = [j for j, v in enumerate(row3_vals) if _norm(v) == _norm('проект')]
    first_block = min(starts) if starts else ws.max_column
    BC = _detect_base_cols([_norm_match(v) for v in row3_vals], first_block)
    if 'inn' not in BC:
        wb.close()
        raise ValueError("Лист «II-Q»: не нашёл колонку «ИНН» в строке 3 — формат не распознан.")

    def base(r, key):
        j = BC.get(key)
        return ws.cell(row=r, column=j + 1).value if j is not None else None

    result = {}
    # Данные начинаются сразу после шапки (стр.3). Строка ИТОГО (ИНН '-' или 7777) отсеется.
    for r in range(4, ws.max_row + 1):
        inn_v = base(r, 'inn')
        if inn_v is None:
            continue
        try:
            inn = str(int(inn_v))
        except (ValueError, TypeError):
            continue
        if inn in ('7777',):  # строка ИТОГО
            continue

        projects = {}
        for pname, cols in blocks:
            if pname.upper().replace(' ', '').replace('-', '') == 'DATFOIIQ':  # общий итог, не проект
                continue
            rec = {}
            for key, cidx in cols.items():
                val = ws.cell(row=r, column=cidx + 1).value
                rec[key] = (str(val).strip() if key == 'condition' else _num(val))
            # Берём ВСЕ проекты (в т.ч. без плана/факта). Какие показать — решает статус
            # 'Актив'/'Неактив' из листа svod (build_dashboard), а не наличие цифр.
            projects[pname] = rec

        result[inn] = {
            'meta': {
                'name': base(r, 'legal') or '',
                'network': base(r, 'network') or '',
                'n_apt': _num(base(r, 'n_apt')),
                'manager': base(r, 'manager') or '',
                'region': base(r, 'region') or '',
                'district': base(r, 'district') or '',
            },
            'total': {
                'plan_iiq': _num(base(r, 'plan_iiq')),
                'fact_iiq': _num(base(r, 'fact_iiq')),
            },
            'projects': projects,
        }
    wb.close()
    return {'blocks': [b[0] for b in blocks], 'pharmacies': result, 'base_cols': BC}


def load_svod_inactive(path=DEFAULT_PATH):
    """Лист «svod»: имена блоков II-Q, у которых статус 'Неактив' (глобально, как в файле).

    Статус в svod — вписанный вручную текст (не формула), одинаков для всех ИНН. Сопоставление
    «строка svod → блок II-Q» делаем по ФОРМУЛЕ строки (она ссылается на колонку II-Q внутри
    нужного блока) — это надёжнее имён ('Synergy' в svod vs блок 'SYENERGY' в II-Q).

    Возвращает set имён блоков (как в строке 2 II-Q). Если листа/формул нет — пустой set.
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
    except Exception:
        return set()
    try:
        if SVOD_SHEET not in wb.sheetnames or SHEET not in wb.sheetnames:
            return set()
        wq = wb[SHEET]
        row2 = [wq.cell(row=2, column=j + 1).value for j in range(wq.max_column)]
        row3 = [wq.cell(row=3, column=j + 1).value for j in range(wq.max_column)]
        # (start_col_1based, block_name) по возрастанию старта
        blocks = [(j + 1, str(row2[j] or '').strip())
                  for j, v in enumerate(row3) if _norm(v) == _norm('проект')]
        if not blocks:
            return set()
        first_block = blocks[0][0]

        def block_for_col(col):
            name = None
            for s, n in blocks:
                if col >= s:
                    name = n
                else:
                    break
            return name

        ws = wb[SVOD_SHEET]
        # колонка «Статус» — по заголовку в верхних строках
        status_col = None
        for r in range(1, 9):
            for c in range(1, ws.max_column + 1):
                if _norm_match(ws.cell(row=r, column=c).value) == 'статус':
                    status_col = c
                    break
            if status_col:
                break
        if not status_col:
            return set()

        inactive = set()
        for r in range(1, ws.max_row + 1):
            sval = str(ws.cell(row=r, column=status_col).value or '').strip().lower()
            if not sval.startswith('неактив'):
                continue
            # первая ссылка на колонку II-Q внутри блока -> имя блока
            chosen = None
            for c in range(1, ws.max_column + 1):
                f = ws.cell(row=r, column=c).value
                if not isinstance(f, str):
                    continue
                for col_str in re.findall(r"II-Q'?!\$?([A-Z]+)\$?\d+", f):
                    ci = column_index_from_string(col_str)
                    if ci >= first_block:
                        chosen = block_for_col(ci)
                        break
                if chosen:
                    break
            if chosen:
                inactive.add(chosen)
        return inactive
    finally:
        wb.close()


if __name__ == '__main__':
    data = load_plans()
    ph = data['pharmacies']
    print(f"Блоки проектов в матрице ({len(data['blocks'])}): {data['blocks']}")
    print(f"\nАптек в планах: {len(ph)}")
    # Пример: первая аптека с планами по проектам
    sample = next((inn for inn, d in ph.items() if d['projects']), None)
    if sample:
        d = ph[sample]
        print(f"\nПример ИНН {sample}: {d['meta']['name']} | менеджер {d['meta']['manager']} | {d['meta']['region']}")
        print(f"  Итог: план IIQ {d['total']['plan_iiq']:,.0f}, факт IIQ {d['total']['fact_iiq']:,.0f}")
        print("  Проекты с планом:")
        for pn, rec in list(d['projects'].items())[:8]:
            print(f"    {pn:<20} планIIQ={rec.get('plan_iiq',0):>14,.0f}  условие={rec.get('condition','')}")
