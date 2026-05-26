import asyncio
import logging
import os
import pandas as pd
from datetime import datetime
from contextlib import suppress
from aiogram.exceptions import TelegramBadRequest
from dotenv import load_dotenv
from aiogram import Bot, Dispatcher, F, types
from aiogram.filters import Command
from aiogram.types import (
    ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove,
    FSInputFile, InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery,
    WebAppInfo
)
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup

from database import (
    get_pharmacies_by_tg_id, get_user_data, register_user,
    update_user_role, get_all_active_users, upsert_user,
    get_all_pharmacies_extended, create_poll, save_poll_answer,
    get_poll_list, get_poll_stats_full
)
from screenshot import update_inn_in_sheet, take_screenshot
from dashboard_sync import sync_dashboard, sync_dashboard_from_excel
from api import start_api

load_dotenv()
# WARNING вместо INFO: убирает шум от aiogram/aiohttp, оставляет наши print и реальные ошибки
logging.basicConfig(level=logging.WARNING)

BOT_TOKEN = os.getenv('BOT_TOKEN')
FEEDBACK_CHANNEL_ID = os.getenv('FEEDBACK_CHANNEL_ID')
WEB_APP_URL = os.getenv('WEB_APP_URL')

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(storage=MemoryStorage())

# БЛОКИРОВКИ
report_lock = asyncio.Lock()
BROADCAST_ACTIVE = False

TEXTS = {
    'ru': {
        'welcome': "Добро пожаловать!",
        'access_denied': "⛔ У вас нет доступа (Статус: Призрак).\nОбратитесь к администратору.",
        'wait': "⏳ Генерирую отчет...",
        'queue': "⏳ Система занята, вы в очереди. Пожалуйста, подождите...",
        'error': "❌ Ошибка.",
        'get_report': "📊 Получить данные",
        'change_lang': "🇺🇿/🇷🇺 Сменить язык",
        'feedback': "✍ Оставить отзыв",
        'feedback_prompt': "✍ Если у вас есть жалобы или предложения, напишите их здесь, и мы передадим руководству Datfo:",
        'feedback_sent': "✅ Ваш отзыв отправлен! Спасибо.",
        'bcast': "📢 Рассылка",
        'admins': "👤 Админы",
        'pharm_list': "🏥 Список Аптек (Excel)",
        'bcast_unique': "📝 Уникальная рассылка",
        'bcast_data': "📊 Рассылка с данными",
        'bcast_list': "📂 Рассылка по списку (Excel)",
        'bcast_poll': "📊 Рассылка Опрос",
        'poll_results': "📈 Результаты опроса",
        'add_admin': "➕ Добавить Админа",
        'del_admin': "🗑 Удалить Админа",
        'back': "🔙 Назад",
        'enter_inn': "👮‍♂️ Введите ИНН для проверки:",
        'enter_id_add': "Введите TG ID нового админа:",
        'enter_id_del': "Введите TG ID для удаления (станет Призраком):",
        'bcast_prompt': "Пришлите сообщение (текст/фото) для рассылки:",
        'poll_title_prompt': "Введите ВНУТРЕННЕЕ название опроса (для отчетов):",
        'poll_target_select': "Кому отправить опрос?",
        'poll_target_all': "👥 Всем юзерам",
        'poll_target_list': "📂 По списку (Excel)",
        'bcast_confirm': "Подтвердить отправку?",
        'bcast_data_confirm': "⚠ Отправить каждому скриншот ЕГО данных?\nЕсли у юзера несколько аптек, придет несколько отчетов.",
        'bcast_list_file': "📂 Пришлите Excel файл (.xlsx).\nБот возьмет ID из ПЕРВОГО столбца (A).",
        'yes': "✅ ДА",
        'no': "❌ НЕТ",
        'poll_yes_btn': "✅ ДА",
        'poll_no_btn': "❌ НЕТ",
        'vote_accepted': "✅ Ваш голос принят!",
        'busy': "🚫 Другая рассылка уже идет! Подождите.",
        'done': "✅ Рассылка завершена.",
        'generating_excel': "📊 Формирую Excel отчет...",
        'select_poll': "📉 Выберите опрос из списка ниже для выгрузки отчета:",
        'file_accepted': "✅ Файл принят. ID найдены.",
        'open_app': "📱 Открыть приложение",
        'upload_excel': "📥 Загрузить Excel",
        'upload_excel_prompt': "📥 Пришлите .xlsx файл с дашбордами аптек (формат «Свод таб», один лист = одна аптека).\n\nДанные сразу обновятся в Mini App.",
        'upload_excel_processing': "⏳ Парсю Excel и обновляю аптеки...",
        'upload_excel_bad_ext': "❌ Это не .xlsx файл.",
        'upload_excel_too_big': "❌ Файл слишком большой (максимум 20 МБ).",
    },
    'uz': {
        'welcome': "Xush kelibsiz!",
        'access_denied': "⛔ Kirish huquqi yo'q (Ghost).\nAdmin bilan bog'laning.",
        'wait': "⏳ Hisobot tayyorlanmoqda...",
        'queue': "⏳ Tizim band, siz navbatdasiz. Iltimos kuting...",
        'error': "❌ Xatolik.",
        'get_report': "📊 Ma'lumot olish",
        'change_lang': "🇺🇿/🇷🇺 Tilni o'zgartirish",
        'feedback': "✍ Fikr qoldirish",
        'feedback_prompt': "✍ Agar sizda shikoyat yoki takliflar bo'lsa, shu yerda yozing va biz Datfo rahbariyatiga yetkazamiz:",
        'feedback_sent': "✅ Fikringiz yuborildi! Rahmat.",
        'bcast': "📢 Xabar yuborish",
        'admins': "👤 Adminlar",
        'pharm_list': "🏥 Dorixonalar ro'yxati (Excel)",
        'bcast_unique': "📝 Maxsus xabar",
        'bcast_data': "📊 Ma'lumotli xabar",
        'bcast_list': "📂 Ro'yxat bo'yicha (Excel)",
        'bcast_poll': "📊 So'rovnoma yuborish",
        'poll_results': "📈 So'rovnoma natijalari",
        'add_admin': "➕ Admin Qo'shish",
        'del_admin': "🗑 Admin O'chirish",
        'back': "🔙 Orqaga",
        'enter_inn': "👮‍♂️ Tekshirish uchun INN kiriting:",
        'enter_id_add': "Yangi adminning TG ID raqamini kiriting:",
        'enter_id_del': "O'chirish uchun TG ID kiriting (Ghost bo'ladi):",
        'bcast_prompt': "Xabar yuboring (matn/rasm):",
        'poll_title_prompt': "So'rovnoma uchun ICHKI nomni kiriting (hisobot uchun):",
        'poll_target_select': "Kimga yuborilsin?",
        'poll_target_all': "👥 Hamma foydalanuvchilarga",
        'poll_target_list': "📂 Ro'yxat bo'yicha (Excel)",
        'bcast_confirm': "Yuborish tasdiqlansinmi?",
        'bcast_data_confirm': "⚠ Har kimga O'Z ma'lumotlarini yuborasizmi?\nAgar bir necha dorixona bo'lsa, bir necha hisobot keladi.",
        'bcast_list_file': "📂 Excel faylni yuboring (.xlsx).\nBot ID larni BIRINCHI ustun (A) dan oladi.",
        'yes': "✅ HA",
        'no': "❌ YO'Q",
        'poll_yes_btn': "✅ HA",
        'poll_no_btn': "❌ YO'Q",
        'vote_accepted': "✅ Ovozingiz qabul qilindi!",
        'busy': "🚫 Hozir boshqa xabar yuborilmoqda! Kuting.",
        'done': "✅ Yakunlandi.",
        'generating_excel': "📊 Excel hisobot yaratilmoqda...",
        'select_poll': "📉 Hisobot olish uchun so'rovnomani tanlang:",
        'file_accepted': "✅ Fayl qabul qilindi. IDlar topildi.",
        'open_app': "📱 Ilovani ochish",
        'upload_excel': "📥 Excel yuklash",
        'upload_excel_prompt': "📥 Dorixonalar dashboardlari bilan .xlsx faylni yuboring («Свод таб» formati, bir varaq = bir dorixona).\n\nMa'lumotlar darhol Mini App ga yangilanadi.",
        'upload_excel_processing': "⏳ Excel tahlil qilinmoqda va dorixonalar yangilanmoqda...",
        'upload_excel_bad_ext': "❌ Bu .xlsx fayl emas.",
        'upload_excel_too_big': "❌ Fayl juda katta (maksimal 20 MB).",
    }
}


def kb_webapp(tg_id, lang):
    """Inline-кнопка под рассылками, открывает Web App с tg_id в URL."""
    if not WEB_APP_URL:
        return None
    url = f"{WEB_APP_URL}?tg_id={tg_id}"
    return InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text=TEXTS[lang]['open_app'], web_app=WebAppInfo(url=url))
    ]])


class AdminState(StatesGroup):
    waiting_for_inn_report = State()
    waiting_for_admin_add = State()
    waiting_for_admin_del = State()
    waiting_for_poll_selection = State()
    waiting_for_dashboard_excel = State()


class FeedbackState(StatesGroup):
    waiting_for_text = State()


class BroadcastState(StatesGroup):
    waiting_for_content = State()
    confirm_unique = State()
    confirm_data = State()
    waiting_for_list_file = State()
    waiting_for_list_content = State()
    confirm_list = State()
    # Poll states ordered correctly
    waiting_for_poll_target_type = State()  # 1. Выбор типа (Все или Список)
    waiting_for_poll_list_file = State()  # 2. Загрузка файла (если список)
    waiting_for_poll_title = State()  # 3. Название
    waiting_for_poll_content = State()  # 4. Контент
    confirm_poll = State()  # 5. Подтверждение


def kb_main(lang, role):
    t = TEXTS[lang]
    btns = [
        [KeyboardButton(text=t['get_report'])],
        [KeyboardButton(text=t['feedback']), KeyboardButton(text=t['change_lang'])]
    ]
    if role in ['admin', 'superadmin']:
        btns.append([KeyboardButton(text=t['bcast']), KeyboardButton(text=t['pharm_list'])])
    if role == 'superadmin':
        btns.append([KeyboardButton(text=t['admins']), KeyboardButton(text=t['upload_excel'])])
    return ReplyKeyboardMarkup(keyboard=btns, resize_keyboard=True)


def kb_back(lang):
    return ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text=TEXTS[lang]['back'])]], resize_keyboard=True)


def kb_yes_no(lang):
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=TEXTS[lang]['yes']), KeyboardButton(text=TEXTS[lang]['no'])]],
        resize_keyboard=True)


# Инлайн клавиатура для опроса (yes/no + WebApp снизу)
def kb_poll_inline(poll_id, lang, tg_id):
    t = TEXTS[lang]
    rows = [[
        InlineKeyboardButton(text=t['poll_yes_btn'], callback_data=f"poll:{poll_id}:yes"),
        InlineKeyboardButton(text=t['poll_no_btn'], callback_data=f"poll:{poll_id}:no")
    ]]
    if WEB_APP_URL:
        url = f"{WEB_APP_URL}?tg_id={tg_id}"
        rows.append([InlineKeyboardButton(text=t['open_app'], web_app=WebAppInfo(url=url))])
    return InlineKeyboardMarkup(inline_keyboard=rows)


# --- EXCEL REPORT GENERATOR ---

async def create_colored_report(results, filename, status_col_index, is_poll=False):
    loop = asyncio.get_running_loop()
    await loop.run_in_executor(None, _blocking_create_report, results, filename, status_col_index, is_poll)


def _blocking_create_report(results, filename, status_col_index, is_poll):
    df = pd.DataFrame(results)

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Report')
        worksheet = writer.sheets['Report']
        max_row = len(df) + 1

        from openpyxl.styles import PatternFill
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        for row in range(2, max_row + 1):
            cell = worksheet.cell(row=row, column=status_col_index)
            val = str(cell.value).lower().strip()

            if is_poll:
                if val == 'yes' or val == 'ha' or val == 'да':
                    cell.fill = green_fill
                elif val == 'no' or val == "yo'q" or val == 'нет':
                    cell.fill = red_fill
                else:
                    cell.fill = gray_fill
            else:
                if "не в боте" in val or "не доставлено" in val or "ошибка" in val or "пропущено" in val:
                    cell.fill = red_fill
                elif "в боте" in val or "доставлено" in val:
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill

        for column_cells in worksheet.columns:
            try:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2
            except:
                pass

@dp.message(Command("myid"))
async def get_my_id(message: types.Message):
    await message.answer(f"Ваш chat_id: {message.chat.id}")


# --- START ---
@dp.message(Command("start"))
async def cmd_start(message: types.Message, state: FSMContext):
    if state: await state.clear()

    tg_id = message.from_user.id
    await register_user(tg_id)

    user = await get_user_data(tg_id)
    role = user['role']
    lang = user['language']

    pharmacies = await get_pharmacies_by_tg_id(tg_id)
    if pharmacies and role == 'ghost':
        await update_user_role(tg_id, 'user')
        role = 'user'
        user = await get_user_data(tg_id)

    if role == 'ghost':
        await message.answer(TEXTS[lang]['access_denied'], reply_markup=ReplyKeyboardRemove())
        return

    admin_text = "\n👮‍♂️ ADMIN" if role in ['admin', 'superadmin'] else ""
    await message.answer(f"{TEXTS[lang]['welcome']}{admin_text}", reply_markup=kb_main(lang, role))


@dp.message(F.text.in_(["🔙 Назад", "🔙 Orqaga"]))
async def go_back(message: types.Message, state: FSMContext):
    await cmd_start(message, state)


@dp.message(F.text.in_(["🇺🇿/🇷🇺 Сменить язык", "🇺🇿/🇷🇺 Tilni o'zgartirish"]))
async def change_lang_req(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    if user['role'] == 'ghost': return await message.answer("⛔ Access Denied")
    await message.answer("Выберите язык / Tilni tanlang:", reply_markup=ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="🇷🇺 Русский"), KeyboardButton(text="🇺🇿 O'zbekcha")]], resize_keyboard=True))


@dp.message(F.text.in_({"🇷🇺 Русский", "🇺🇿 O'zbekcha"}))
async def set_lang(message: types.Message, state: FSMContext):
    lang = 'ru' if "Русский" in message.text else 'uz'
    user = await get_user_data(message.from_user.id)
    if user['role'] == 'ghost': return await message.answer("⛔ Access Denied")
    await upsert_user(message.from_user.id, user['role'], lang)
    await cmd_start(message, state)


# --- ОТЗЫВЫ (Feature 2) ---
@dp.message(F.text.in_(["✍ Оставить отзыв", "✍ Fikr qoldirish"]))
async def feedback_start(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    lang = user['language']
    if user['role'] == 'ghost': return

    await message.answer(TEXTS[lang]['feedback_prompt'], reply_markup=kb_back(lang))
    await state.set_state(FeedbackState.waiting_for_text)


@dp.message(FeedbackState.waiting_for_text)
async def feedback_send(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    lang = user['language']
    tg_id = message.from_user.id

    pharms = await get_pharmacies_by_tg_id(tg_id)
    pharm_names = ", ".join([p.get('pharmacy_name', 'Unknown') or 'Unknown' for p in pharms])
    inns = ", ".join([p.get('inn', 'No INN') for p in pharms])

    report_text = (
        f"📣 <b>НОВЫЙ ОТЗЫВ</b>\n"
        f"👤 <b>ID:</b> <code>{tg_id}</code>\n"
        f"🏥 <b>Аптека:</b> {pharm_names}\n"
        f"🔢 <b>ИНН:</b> {inns}\n"
        f"➖➖➖➖➖➖➖\n"
        f"{message.text}"
    )

    if FEEDBACK_CHANNEL_ID:
        try:
            await bot.send_message(FEEDBACK_CHANNEL_ID, report_text, parse_mode='HTML')
            await message.answer(TEXTS[lang]['feedback_sent'])
        except Exception as e:
            await message.answer(f"Error sending feedback: {e}")
    else:
        await message.answer("Error: FEEDBACK_CHANNEL_ID not set in .env")

    await cmd_start(message, state)


# --- ПОЛУЧЕНИЕ ОТЧЕТА ---
@dp.message(F.text.in_(["📊 Получить данные", "📊 Ma'lumot olish"]))
async def report_req(message: types.Message, state: FSMContext):
    tg_id = message.from_user.id
    user = await get_user_data(tg_id)
    lang = user['language']

    if user['role'] == 'ghost': return await message.answer(TEXTS[lang]['access_denied'])

    if user['role'] in ['admin', 'superadmin']:
        await message.answer(TEXTS[lang]['enter_inn'], reply_markup=kb_back(lang))
        await state.set_state(AdminState.waiting_for_inn_report)
    else:
        pharmacies = await get_pharmacies_by_tg_id(tg_id)
        if not pharmacies: return await message.answer(TEXTS[lang]['error'])
        for pharm in pharmacies:
            await generate_report(message, pharm['inn'], lang)


@dp.message(AdminState.waiting_for_inn_report)
async def report_admin_exec(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    if user['role'] == 'ghost': return
    if not message.text.isdigit(): return await message.answer("INN = 0-9")
    await generate_report(message, message.text, user['language'])
    await state.clear()
    await cmd_start(message, state)


async def generate_report(message, inn, lang):
    if report_lock.locked(): await message.answer(TEXTS[lang]['queue'])
    msg = await message.answer(f"{TEXTS[lang]['wait']} (INN: {inn})")
    async with report_lock:
        try:
            res = await update_inn_in_sheet(inn)
            if not res: raise Exception("Google Error")
            fname = f"rep_{message.from_user.id}_{inn}.png"
            await take_screenshot(output_filename=fname)
            await message.answer_photo(types.FSInputFile(fname))
            await msg.delete()
            if os.path.exists(fname): os.remove(fname)
        except Exception:
            await msg.edit_text(f"{TEXTS[lang]['error']} (INN: {inn})")


# --- ВЫГРУЗКА СПИСКА АПТЕК ---
@dp.message(F.text.in_(["🏥 Список Аптек (Excel)", "🏥 Dorixonalar ro'yxati (Excel)"]))
async def get_pharmacy_list_excel(message: types.Message):
    user = await get_user_data(message.from_user.id)
    if user['role'] not in ['admin', 'superadmin']: return

    msg = await message.answer("⏳ Формирую файл...")

    rows = await get_all_pharmacies_extended()
    data = []
    for r in rows:
        role = r['role']
        status = "В боте" if role and role != 'ghost' else "Не в боте"

        data.append({
            'ID': r['id'],
            'INN': r['inn'],
            'TG_ID': r['owner_tg_id'],
            'Business': r['business_name'],
            'Pharmacy': r['pharmacy_name'],
            'Status': status
        })

    if not data: return await msg.edit_text("❌ База пуста.")

    filename = "pharmacies_list.xlsx"
    await create_colored_report(data, filename, status_col_index=6)

    await message.answer_document(FSInputFile(filename), caption="Список всех аптек из базы (Зеленый = В боте)")
    await msg.delete()
    if os.path.exists(filename): os.remove(filename)


# --- МЕНЮ РАССЫЛКИ ---
@dp.message(F.text.in_(["📢 Рассылка", "📢 Xabar yuborish"]))
async def bcast_menu(message: types.Message):
    user = await get_user_data(message.from_user.id)
    if user['role'] not in ['admin', 'superadmin']: return
    lang = user['language']

    kb = ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text=TEXTS[lang]['bcast_unique']), KeyboardButton(text=TEXTS[lang]['bcast_data'])],
        [KeyboardButton(text=TEXTS[lang]['bcast_poll'])],
        [KeyboardButton(text=TEXTS[lang]['bcast_list']), KeyboardButton(text=TEXTS[lang]['poll_results'])],
        [KeyboardButton(text=TEXTS[lang]['back'])]
    ], resize_keyboard=True)
    await message.answer(TEXTS[lang]['bcast'], reply_markup=kb)


# --- 1. УНИКАЛЬНАЯ РАССЫЛКА ---
@dp.message(F.text.in_(["📝 Уникальная рассылка", "📝 Maxsus xabar"]))
async def bcast_unique_start(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    if user['role'] not in ['admin', 'superadmin']: return
    lang = user['language']

    global BROADCAST_ACTIVE
    if BROADCAST_ACTIVE: return await message.answer(TEXTS[lang]['busy'])
    await message.answer(TEXTS[lang]['bcast_prompt'], reply_markup=kb_back(lang))
    await state.set_state(BroadcastState.waiting_for_content)


@dp.message(BroadcastState.waiting_for_content)
async def bcast_unique_content(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    lang = user['language']
    await state.update_data(mid=message.message_id, cid=message.chat.id)
    await message.answer(TEXTS[lang]['bcast_confirm'], reply_markup=kb_yes_no(lang))
    await state.set_state(BroadcastState.confirm_unique)


@dp.message(BroadcastState.confirm_unique)
async def bcast_unique_send(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    lang = user['language']
    if message.text not in [TEXTS['ru']['yes'], TEXTS['uz']['yes']]:
        await state.clear()
        return await cmd_start(message, state)

    global BROADCAST_ACTIVE
    if BROADCAST_ACTIVE: return await message.answer(TEXTS[lang]['busy'])
    BROADCAST_ACTIVE = True

    try:
        data = await state.get_data()
        users = await get_all_active_users()
        st = await message.answer(f"🚀 Start! Users: {len(users)}", reply_markup=ReplyKeyboardRemove())

        results = []
        for u in users:
            tg_id = u['telegram_id']
            pharms = await get_pharmacies_by_tg_id(tg_id)
            inns_str = ", ".join([p['inn'] for p in pharms]) if pharms else "No INN"

            status = "Не доставлено"
            try:
                await bot.copy_message(
                    tg_id, data['cid'], data['mid'],
                    reply_markup=kb_webapp(tg_id, u['language'])
                )
                status = "Доставлено"
                await asyncio.sleep(0.05)
            except Exception as e:
                status = f"Ошибка: {str(e)}"

            results.append({'TG_ID': tg_id, 'INNs': inns_str, 'Status': status})

        await message.answer(TEXTS[lang]['done'] + f" ({len(users)})")
        await st.delete()

        await message.answer(TEXTS[lang]['generating_excel'])
        fname = "unique_report.xlsx"
        await create_colored_report(results, fname, 3)
        await message.answer_document(FSInputFile(fname), caption="Отчет по рассылке")
        if os.path.exists(fname): os.remove(fname)

    except Exception as e:
        await message.answer(f"Error: {e}")
    finally:
        BROADCAST_ACTIVE = False
        await state.clear()
        await cmd_start(message, state)


# --- 2. РАССЫЛКА С ДАННЫМИ ---
@dp.message(F.text.in_(["📊 Рассылка с данными", "📊 Ma'lumotli xabar"]))
async def bcast_data_start(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    if user['role'] not in ['admin', 'superadmin']: return
    lang = user['language']

    global BROADCAST_ACTIVE
    if BROADCAST_ACTIVE: return await message.answer(TEXTS[lang]['busy'])
    await message.answer(TEXTS[lang]['bcast_data_confirm'], reply_markup=kb_yes_no(lang))
    await state.set_state(BroadcastState.confirm_data)


@dp.message(BroadcastState.confirm_data)
async def bcast_data_send(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    lang = user['language']
    if message.text not in [TEXTS['ru']['yes'], TEXTS['uz']['yes']]:
        await state.clear()
        return await cmd_start(message, state)
    global BROADCAST_ACTIVE
    if BROADCAST_ACTIVE: return await message.answer(TEXTS[lang]['busy'])
    BROADCAST_ACTIVE = True

    try:
        users = await get_all_active_users()
        st = await message.answer(f"🚀 Generating reports...", reply_markup=ReplyKeyboardRemove())

        results = []
        async with report_lock:
            for i, u in enumerate(users):
                tg_id = u['telegram_id']
                pharms = await get_pharmacies_by_tg_id(tg_id)

                if not pharms:
                    results.append({'TG_ID': tg_id, 'INN': 'No INN', 'Status': 'Пропущено (нет аптеки)'})
                    continue

                for p in pharms:
                    inn = p['inn']
                    status = "Не доставлено"
                    try:
                        await update_inn_in_sheet(inn)
                        fname = f"bc_{tg_id}_{inn}.png"
                        await take_screenshot(fname)
                        caption = "Ваш отчет" if u['language'] == 'ru' else "Sizning hisobotingiz"
                        await bot.send_photo(
                            tg_id, types.FSInputFile(fname), caption=caption,
                            reply_markup=kb_webapp(tg_id, u['language'])
                        )
                        status = "Доставлено"
                        if os.path.exists(fname): os.remove(fname)
                    except Exception as e:
                        status = f"Ошибка: {str(e)}"

                    results.append({'TG_ID': tg_id, 'INN': inn, 'Status': status})

        await message.answer(TEXTS[lang]['done'])
        await st.delete()

        await message.answer(TEXTS[lang]['generating_excel'])
        fname = "data_report.xlsx"
        await create_colored_report(results, fname, 3)
        await message.answer_document(FSInputFile(fname), caption="Отчет по рассылке с данными")
        if os.path.exists(fname): os.remove(fname)

    except Exception as e:
        print(f"Data Broadcast Error: {e}")
    finally:
        BROADCAST_ACTIVE = False
        await state.clear()
        await cmd_start(message, state)


# --- 3. РАССЫЛКА ПО СПИСКУ ---
@dp.message(F.text.in_(["📂 Рассылка по списку (Excel)", "📂 Ro'yxat bo'yicha (Excel)"]))
async def bcast_list_start(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    if user['role'] not in ['admin', 'superadmin']: return
    lang = user['language']
    global BROADCAST_ACTIVE
    if BROADCAST_ACTIVE: return await message.answer(TEXTS[lang]['busy'])

    await message.answer(TEXTS[lang]['bcast_list_file'], reply_markup=kb_back(lang))
    await state.set_state(BroadcastState.waiting_for_list_file)


@dp.message(BroadcastState.waiting_for_list_file)
async def bcast_list_file_handler(message: types.Message, state: FSMContext):
    if not message.document:
        return await message.answer("❌ Жду файл .xlsx")

    file_id = message.document.file_id
    file = await bot.get_file(file_id)
    file_path = f"upload_{message.from_user.id}.xlsx"
    await bot.download_file(file.file_path, file_path)

    try:
        df = pd.read_excel(file_path)
        target_ids = df.iloc[:, 0].dropna().unique().tolist()
        clean_ids = []
        for tid in target_ids:
            try:
                clean_ids.append(int(tid))
            except:
                pass

        if not clean_ids:
            return await message.answer("❌ В файле не найдено ID (Столбец A).")

        await state.update_data(target_ids=clean_ids)
        if os.path.exists(file_path): os.remove(file_path)

        user = await get_user_data(message.from_user.id)
        lang = user['language']
        await message.answer(f"✅ Файл принят. Найдено ID: {len(clean_ids)}.\n{TEXTS[lang]['bcast_prompt']}")
        await state.set_state(BroadcastState.waiting_for_list_content)

    except Exception as e:
        await message.answer(f"❌ Ошибка чтения файла: {e}")
        if os.path.exists(file_path): os.remove(file_path)


@dp.message(BroadcastState.waiting_for_list_content)
async def bcast_list_content(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    lang = user['language']
    await state.update_data(mid=message.message_id, cid=message.chat.id)
    await message.answer(TEXTS[lang]['bcast_confirm'], reply_markup=kb_yes_no(lang))
    await state.set_state(BroadcastState.confirm_list)


@dp.message(BroadcastState.confirm_list)
async def bcast_list_send(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    lang = user['language']
    if message.text not in [TEXTS['ru']['yes'], TEXTS['uz']['yes']]:
        await state.clear()
        return await cmd_start(message, state)

    global BROADCAST_ACTIVE
    if BROADCAST_ACTIVE: return await message.answer(TEXTS[lang]['busy'])
    BROADCAST_ACTIVE = True

    try:
        data = await state.get_data()
        targets = data['target_ids']
        st = await message.answer(f"🚀 Start! Targets: {len(targets)}", reply_markup=ReplyKeyboardRemove())

        results = []
        for tg_id in targets:
            pharms = await get_pharmacies_by_tg_id(tg_id)
            inns_str = ", ".join([p['inn'] for p in pharms]) if pharms else "No INN"
            status = "Не доставлено"
            try:
                target_user = await get_user_data(tg_id)
                target_lang = target_user['language'] if target_user else 'ru'
                await bot.copy_message(
                    tg_id, data['cid'], data['mid'],
                    reply_markup=kb_webapp(tg_id, target_lang)
                )
                status = "Доставлено"
                await asyncio.sleep(0.05)
            except Exception as e:
                status = f"Ошибка: {str(e)}"

            results.append({'TG_ID': tg_id, 'INNs': inns_str, 'Status': status})

        await message.answer(TEXTS[lang]['done'])
        await st.delete()

        await message.answer(TEXTS[lang]['generating_excel'])
        fname = "list_report.xlsx"
        await create_colored_report(results, fname, 3)
        await message.answer_document(FSInputFile(fname), caption="Отчет по рассылке (по списку)")
        if os.path.exists(fname): os.remove(fname)

    except Exception as e:
        await message.answer(f"Error: {e}")
    finally:
        BROADCAST_ACTIVE = False
        await state.clear()
        await cmd_start(message, state)


# --- 4. РАССЫЛКА ОПРОС ---

@dp.message(F.text.in_(["📊 Рассылка Опрос", "📊 So'rovnoma yuborish"]))
async def bcast_poll_start(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    if user['role'] not in ['admin', 'superadmin']: return
    lang = user['language']

    global BROADCAST_ACTIVE
    if BROADCAST_ACTIVE: return await message.answer(TEXTS[lang]['busy'])

    # ШАГ 1: СПРАШИВАЕМ КОМУ ОТПРАВИТЬ
    kb = ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text=TEXTS[lang]['poll_target_all'])],
        [KeyboardButton(text=TEXTS[lang]['poll_target_list'])],
        [KeyboardButton(text=TEXTS[lang]['back'])]
    ], resize_keyboard=True)

    await message.answer(TEXTS[lang]['poll_target_select'], reply_markup=kb)
    await state.set_state(BroadcastState.waiting_for_poll_target_type)


@dp.message(BroadcastState.waiting_for_poll_target_type)
async def bcast_poll_target_choice(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    lang = user['language']

    if message.text == TEXTS[lang]['poll_target_all']:
        # Если ВСЕМ -> сразу спрашиваем НАЗВАНИЕ
        await state.update_data(poll_target_type='all')
        await message.answer(TEXTS[lang]['poll_title_prompt'], reply_markup=kb_back(lang))
        await state.set_state(BroadcastState.waiting_for_poll_title)

    elif message.text == TEXTS[lang]['poll_target_list']:
        # Если СПИСОК -> сначала просим ФАЙЛ
        await state.update_data(poll_target_type='list')
        await message.answer(TEXTS[lang]['bcast_list_file'], reply_markup=kb_back(lang))
        await state.set_state(BroadcastState.waiting_for_poll_list_file)

    elif message.text == TEXTS[lang]['back']:
        await cmd_start(message, state)
    else:
        await message.answer("Please select an option from the keyboard.")


@dp.message(BroadcastState.waiting_for_poll_list_file)
async def bcast_poll_list_file_handler(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    lang = user['language']

    if not message.document:
        return await message.answer("❌ Жду файл .xlsx")

    file_id = message.document.file_id
    file = await bot.get_file(file_id)
    file_path = f"upload_poll_{message.from_user.id}.xlsx"
    await bot.download_file(file.file_path, file_path)

    try:
        df = pd.read_excel(file_path)
        target_ids = df.iloc[:, 0].dropna().unique().tolist()
        clean_ids = []
        for tid in target_ids:
            try:
                clean_ids.append(int(tid))
            except:
                pass

        if not clean_ids:
            return await message.answer("❌ В файле не найдено ID (Столбец A).")

        # Сохраняем ID для рассылки
        await state.update_data(poll_target_ids=clean_ids)
        if os.path.exists(file_path): os.remove(file_path)

        # ФАЙЛ ПРИНЯТ -> ТЕПЕРЬ СПРАШИВАЕМ НАЗВАНИЕ
        await message.answer(f"{TEXTS[lang]['file_accepted']}\n{TEXTS[lang]['poll_title_prompt']}")
        await state.set_state(BroadcastState.waiting_for_poll_title)

    except Exception as e:
        await message.answer(f"❌ Ошибка чтения файла: {e}")
        if os.path.exists(file_path): os.remove(file_path)


@dp.message(BroadcastState.waiting_for_poll_title)
async def bcast_poll_title(message: types.Message, state: FSMContext):
    # НАЗВАНИЕ ПРИНЯТО -> ПРОСИМ КОНТЕНТ (КАРТИНКА/ТЕКСТ)
    await state.update_data(poll_title=message.text)
    user = await get_user_data(message.from_user.id)
    lang = user['language']
    await message.answer(TEXTS[lang]['bcast_prompt'])
    await state.set_state(BroadcastState.waiting_for_poll_content)


@dp.message(BroadcastState.waiting_for_poll_content)
async def bcast_poll_content(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    lang = user['language']
    # КОНТЕНТ ПРИНЯТ -> ПОДТВЕРЖДЕНИЕ
    await state.update_data(mid=message.message_id, cid=message.chat.id)
    await message.answer(TEXTS[lang]['bcast_confirm'], reply_markup=kb_yes_no(lang))
    await state.set_state(BroadcastState.confirm_poll)


@dp.message(BroadcastState.confirm_poll)
async def bcast_poll_send(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    lang = user['language']
    if message.text not in [TEXTS['ru']['yes'], TEXTS['uz']['yes']]:
        await state.clear()
        return await cmd_start(message, state)

    global BROADCAST_ACTIVE
    if BROADCAST_ACTIVE: return await message.answer(TEXTS[lang]['busy'])
    BROADCAST_ACTIVE = True

    try:
        data = await state.get_data()
        poll_title = data['poll_title']
        target_type = data.get('poll_target_type', 'all')

        # 1. Создаем опрос в БД
        poll_id = await create_poll(poll_title)

        # 2. Определяем список получателей
        all_active_users = await get_all_active_users()  # [{telegram_id, language, role}, ...]

        final_targets = []

        if target_type == 'all':
            final_targets = all_active_users
        else:
            # Если по списку, фильтруем всех активных юзеров, оставляем только тех, кто есть в файле
            list_ids = data.get('poll_target_ids', [])
            final_targets = [u for u in all_active_users if u['telegram_id'] in list_ids]

        st = await message.answer(f"🚀 Start Poll Broadcast ({target_type})! Users: {len(final_targets)}",
                                  reply_markup=ReplyKeyboardRemove())

        # 3. Рассылаем
        count = 0
        for u in final_targets:
            tg_id = u['telegram_id']
            u_lang = u['language']

            try:
                await bot.copy_message(
                    chat_id=tg_id,
                    from_chat_id=data['cid'],
                    message_id=data['mid'],
                    reply_markup=kb_poll_inline(poll_id, u_lang, tg_id)
                )
                count += 1
                await asyncio.sleep(0.05)
            except Exception as e:
                pass

        await message.answer(f"{TEXTS[lang]['done']}. Отправлено: {count}")
        await st.delete()

    except Exception as e:
        await message.answer(f"Error: {e}")
    finally:
        BROADCAST_ACTIVE = False
        await state.clear()
        await cmd_start(message, state)


# --- ОБРАБОТКА НАЖАТИЯ НА КНОПКИ ОПРОСА (UPDATED) ---
@dp.callback_query(F.data.startswith('poll:'))
async def poll_vote_handler(callback: CallbackQuery):
    # data format: poll:123:yes
    parts = callback.data.split(':')
    poll_id = int(parts[1])
    answer = parts[2]  # yes / no
    tg_id = callback.from_user.id

    # Сохраняем в БД (перезаписываем, если уже голосовал)
    await save_poll_answer(poll_id, tg_id, answer)

    # Получаем язык юзера для ответа
    user = await get_user_data(tg_id)
    lang = user['language'] if user else 'ru'

    # Формируем текст подтверждения
    date_str = datetime.now().strftime("%H:%M %d.%m.%Y")

    # Разделитель, чтобы не дублировать текст при повторном нажатии
    separator = "\n\n➖➖➖➖➖➖➖➖\n"

    if lang == 'ru':
        ans_text = "ДА" if answer == 'yes' else "НЕТ"
        confirm_text = f"{separator}✅ Вы выбрали: {ans_text}\n📅 Дата: {date_str}"
    else:
        ans_text = "HA" if answer == 'yes' else "YO'Q"
        confirm_text = f"{separator}✅ Siz tanladingiz: {ans_text}\n📅 Sana: {date_str}"

    # Получаем текущую клавиатуру, чтобы она не исчезла
    current_keyboard = kb_poll_inline(poll_id, lang, tg_id)

    # Редактируем сообщение: очищаем старый статус (если был) и добавляем новый
    with suppress(TelegramBadRequest):  # Игнорируем ошибку, если текст не изменился (тот же выбор)
        if callback.message.caption:
            original_caption = callback.message.caption.split(separator)[0]  # Берем только исходный текст
            await callback.message.edit_caption(
                caption=original_caption + confirm_text,
                reply_markup=current_keyboard
            )
        elif callback.message.text:
            original_text = callback.message.text.split(separator)[0]  # Берем только исходный текст
            await callback.message.edit_text(
                text=original_text + confirm_text,
                reply_markup=current_keyboard
            )

    await callback.answer(TEXTS[lang]['vote_accepted'])


# --- ВЫГРУЗКА РЕЗУЛЬТАТОВ ОПРОСА ---
@dp.message(F.text.in_(["📈 Результаты опроса", "📈 So'rovnoma natijalari"]))
async def poll_results_menu(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    if user['role'] not in ['admin', 'superadmin']: return
    lang = user['language']

    # Получаем список последних опросов
    polls = await get_poll_list(limit=15)
    if not polls:
        return await message.answer("Опросов пока нет.")

    # Создаем клавиатуру с названиями опросов
    btns = []
    for p in polls:
        # Формат кнопки: "ID. Title"
        btns.append([KeyboardButton(text=f"Poll #{p['id']}: {p['title']}")])
    btns.append([KeyboardButton(text=TEXTS[lang]['back'])])

    await message.answer(TEXTS[lang]['select_poll'],
                         reply_markup=ReplyKeyboardMarkup(keyboard=btns, resize_keyboard=True))
    await state.set_state(AdminState.waiting_for_poll_selection)


@dp.message(AdminState.waiting_for_poll_selection)
async def poll_results_generate(message: types.Message, state: FSMContext):
    if "Poll #" not in message.text:
        await state.clear()
        return await cmd_start(message, state)

    try:
        # Парсим ID из текста кнопки "Poll #123: Title"
        poll_id_str = message.text.split(':')[0].replace("Poll #", "")
        poll_id = int(poll_id_str)

        msg = await message.answer("⏳ Собираю ответы...")

        # Получаем данные (Active users + Answers + Pharmacy info)
        rows = await get_poll_stats_full(poll_id)

        data = []
        for r in rows:
            # Нормализация ответа
            raw_ans = r['answer']
            if raw_ans == 'yes':
                ans_text = "YES"
            elif raw_ans == 'no':
                ans_text = "NO"
            else:
                ans_text = "NO ANSWER"

            data.append({
                'TG_ID': r['telegram_id'],
                'INN': r['inn'] or 'No INN',
                'Pharmacy': r['pharmacy_name'] or 'Unknown',
                'Vote': ans_text
            })

        filename = f"poll_{poll_id}_results.xlsx"
        # 4-я колонка (D) это Vote. Индекс 4.
        await create_colored_report(data, filename, status_col_index=4, is_poll=True)

        await message.answer_document(FSInputFile(filename), caption=f"Результаты: {message.text}")
        await msg.delete()
        if os.path.exists(filename): os.remove(filename)

        await state.clear()
        await cmd_start(message, state)

    except Exception as e:
        await message.answer(f"Error: {e}")
        await state.clear()
        await cmd_start(message, state)


# --- УПРАВЛЕНИЕ АДМИНАМИ ---
@dp.message(F.text.in_(["👤 Админы", "👤 Adminlar"]))
async def admin_manage(message: types.Message):
    user = await get_user_data(message.from_user.id)
    if user['role'] != 'superadmin': return
    lang = user['language']
    kb = ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text=TEXTS[lang]['add_admin']), KeyboardButton(text=TEXTS[lang]['del_admin'])],
        [KeyboardButton(text=TEXTS[lang]['back'])]
    ], resize_keyboard=True)
    await message.answer(TEXTS[lang]['admins'], reply_markup=kb)


@dp.message(F.text.in_(["➕ Добавить Админа", "➕ Admin Qo'shish"]))
async def add_admin(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    if user['role'] != 'superadmin': return
    lang = user['language']
    await message.answer(TEXTS[lang]['enter_id_add'], reply_markup=kb_back(lang))
    await state.set_state(AdminState.waiting_for_admin_add)


@dp.message(AdminState.waiting_for_admin_add)
async def add_admin_ex(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    if user['role'] != 'superadmin': return
    if not message.text.isdigit(): return
    tid = int(message.text)
    await register_user(tid)
    await update_user_role(tid, 'admin')
    await message.answer("✅ OK, Admin added.")
    await state.clear()
    await cmd_start(message, state)


@dp.message(F.text.in_(["🗑 Удалить Админа", "🗑 Admin O'chirish"]))
async def del_admin(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    if user['role'] != 'superadmin': return
    lang = user['language']
    await message.answer(TEXTS[lang]['enter_id_del'], reply_markup=kb_back(lang))
    await state.set_state(AdminState.waiting_for_admin_del)


@dp.message(AdminState.waiting_for_admin_del)
async def del_admin_ex(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    if user['role'] != 'superadmin': return
    if not message.text.isdigit(): return
    tid = int(message.text)
    await update_user_role(tid, 'ghost')
    await message.answer("✅ OK, Admin removed (ghost).")
    await state.clear()
    await cmd_start(message, state)


# --- ЗАГРУЗКА EXCEL С ДАШБОРДАМИ (superadmin) ---
MAX_DASHBOARD_XLSX_BYTES = 20 * 1024 * 1024  # 20 MB

@dp.message(F.text.in_(["📥 Загрузить Excel", "📥 Excel yuklash"]))
async def upload_dashboard_excel_start(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    if user['role'] != 'superadmin': return
    lang = user['language']
    await message.answer(TEXTS[lang]['upload_excel_prompt'], reply_markup=kb_back(lang))
    await state.set_state(AdminState.waiting_for_dashboard_excel)


@dp.message(AdminState.waiting_for_dashboard_excel)
async def upload_dashboard_excel_recv(message: types.Message, state: FSMContext):
    user = await get_user_data(message.from_user.id)
    if user['role'] != 'superadmin':
        await state.clear()
        return

    lang = user['language']
    doc = message.document
    if not doc:
        return await message.answer("❌ Жду документ .xlsx")
    if not (doc.file_name or '').lower().endswith('.xlsx'):
        return await message.answer(TEXTS[lang]['upload_excel_bad_ext'])
    if doc.file_size and doc.file_size > MAX_DASHBOARD_XLSX_BYTES:
        return await message.answer(TEXTS[lang]['upload_excel_too_big'])

    file_path = f"dashboard_upload_{message.from_user.id}.xlsx"
    try:
        file = await bot.get_file(doc.file_id)
        await bot.download_file(file.file_path, file_path)
    except Exception as e:
        await state.clear()
        return await message.answer(f"❌ Ошибка скачивания: {e}")

    status_msg = await message.answer(TEXTS[lang]['upload_excel_processing'])

    try:
        result = await sync_dashboard_from_excel(file_path)
    except Exception as e:
        result = {'error': f'{type(e).__name__}: {e}', 'updated': 0}
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)

    # Формируем отчёт
    if result.get('error'):
        text = f"❌ Ошибка: {result['error']}"
    else:
        text = (
            f"✅ Обновлено: <b>{result['updated']}</b> из {result['total_sheets']} листов\n"
        )
        skipped = result.get('skipped_no_tg') or []
        if skipped:
            preview = ", ".join(s['inn'] for s in skipped[:5])
            more = f" и ещё {len(skipped) - 5}" if len(skipped) > 5 else ""
            text += f"\n⚠️ Без TG_ID ({len(skipped)}): {preview}{more}"
        errors = result.get('errors') or []
        if errors:
            preview = "; ".join(f"{e['inn']}: {e['error']}" for e in errors[:3])
            text += f"\n❌ Ошибок: {len(errors)}\n{preview}"

    await status_msg.edit_text(text, parse_mode='HTML')
    await state.clear()
    await cmd_start(message, state)


# --- ЗАДАЧИ ---
# sync_pharmacies (старый мастер-Data лист) отключён.
# Источник истины — Excel-листы дашборда, sync_dashboard сам upsert-ит аптеки и привязывает к TG_ID.
async def scheduled_tasks():
    await sync_dashboard()
    while True:
        await asyncio.sleep(300)
        await sync_dashboard()


async def main():
    print("🤖 Бот запущен (UPDATED: Poll Buttons Stay)!")
    asyncio.create_task(scheduled_tasks())
    api_port = int(os.getenv('API_PORT', '8080'))
    asyncio.create_task(start_api(port=api_port))
    await bot.delete_webhook(drop_pending_updates=True)
    await dp.start_polling(bot)

try:
    if __name__ == "__main__":
        asyncio.run(main())
except (KeyboardInterrupt, SystemExit):
    print("Bot stopped.")