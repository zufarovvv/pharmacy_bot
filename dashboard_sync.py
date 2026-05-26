"""
Синхронизация dashboard_data из Google-таблицы формата "Свод таб new".

Структура одного листа (= одна аптека):
  R3-4: meta (ИНН в C4, Аптека D4, Регион F4, Район H4, Категория J4, Менеджер L4)
  R5-6: двухуровневые заголовки колонок
  R7+ : строки проектов (№ в B, название в C, дальше план/факт по месяцам/кварталу)
  R??:  строка ОБЩЕЕ (B = "ОБЩЕЕ") — итоги по аптеке

Колонки:
   1: пусто
   2: №
   3: Pharmacy/Project name
   4: ПЛАН I-Q (квартальный план)
   5/6/7:   План/Факт/% за Январь
   8/9/10:  План/Факт/% за Февраль
   11/12/13: План/Факт/% за Март
   14: Условия (Закуп/Продажа)
   15: ВП Квартал (%)
   16: визуальный бар (пропускаем)
   17: Сумма до выполнения плана (или "+ ..." если перевыполнено)
   18: Статус (NEW / Активен)
   19: % бонуса (0.07 = 7%)
   20: Сумма бонуса (заработано)
"""
import asyncio
import os
from collections import defaultdict

import gspread
from dotenv import load_dotenv
from google.oauth2.service_account import Credentials

from database import upsert_pharmacy_full, upsert_pharmacy_dashboard_only

load_dotenv()

SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive',
]
CREDS_FILE = os.getenv('GOOGLE_CREDS_FILE', 'creds.json')
DASHBOARD_SHEET_ID = os.getenv('DASHBOARD_SHEET_ID')

# Имена месяцев в этом же порядке как в листе (jan/feb/mar)
MONTHS_RU = ['january', 'february', 'march']

# Координаты meta-полей (1-based: row, col)
META_FIELDS = {
    'inn':              (4, 3),
    'name':             (4, 4),
    'region':           (4, 6),
    'district':         (4, 8),
    'category':         (4, 10),
    'manager':          (4, 12),
    'manager_phone':    (4, 13),  # M4 — телефон менеджера (например, +998901234567)
    'manager_username': (4, 14),  # N4 — Telegram-юзернейм менеджера (например, @datfo_manager)
}

# Где может лежать TG_ID — пробуем по очереди. Первое найденное число берём.
# Менеджер может вписать в любую из этих ячеек:
TG_ID_CANDIDATE_CELLS = [
    (4, 1),   # A4 — оптимально
    (3, 1),   # A3
    (2, 1),   # A2
    (1, 1),   # A1
    (4, 2),   # B4
    (2, 2),   # B2
    (1, 14),  # N1
    (2, 14),  # N2
]

# Колонки данных проекта (1-based)
COL_NUMBER = 2
COL_NAME = 3
COL_QUARTER_PLAN = 4
COL_MONTHS = [
    # (plan, fact, percent) для январь/февраль/март
    (5, 6, 7),
    (8, 9, 10),
    (11, 12, 13),
]
COL_CONDITION = 14
COL_QUARTER_PERCENT = 15
COL_REMAINING = 17
COL_STATUS = 18
COL_BONUS_PCT = 19
COL_BONUS_AMOUNT = 20

PROJECTS_START_ROW = 7  # с какой строки идут проекты


def _open_workbook():
    creds = Credentials.from_service_account_file(CREDS_FILE, scopes=SCOPES)
    client = gspread.authorize(creds)
    return client.open_by_key(DASHBOARD_SHEET_ID)


def _to_float(val, default=0.0):
    if val is None or val == '':
        return default
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(' ', '').replace(',', '.').replace('+', '').strip()
    try:
        return float(s)
    except ValueError:
        return default


def _to_int_pct(val):
    """Десятичная доля (0.135 = 13.5%, 2.2 = 220%) → целый процент. Excel ВСЕГДА хранит как долю."""
    return int(round(_to_float(val) * 100))


def _fmt_money(val):
    """Число → '10 000 000' (полные цифры с разделителем-пробелом)."""
    f = _to_float(val)
    if f == 0:
        return '0'
    return f'{int(round(f)):,}'.replace(',', ' ')


def _status_for(percent):
    if percent >= 100:
        return 'completed'
    if percent >= 50:
        return 'partial'
    return 'critical'


def _cell(rows, row_idx, col_idx):
    """Безопасный доступ. rows — список списков (0-based). row_idx/col_idx — 1-based."""
    r = row_idx - 1
    c = col_idx - 1
    if r < 0 or r >= len(rows):
        return None
    if c < 0 or c >= len(rows[r]):
        return None
    return rows[r][c]


def _parse_pharmacy_sheet(rows):
    """Парсит одну вкладку. Возвращает dict с данными аптеки или None если лист не подходит."""
    # Проверка: есть ли ИНН в C4?
    inn_raw = _cell(rows, *META_FIELDS['inn'])
    inn_num = _to_float(inn_raw)
    if inn_num <= 0:
        return None
    inn = str(int(inn_num))

    # TG_ID ищем по нескольким возможным ячейкам, берём первое валидное число (≥6 цифр, ≠ ИНН)
    tg_id = None
    for (r, c) in TG_ID_CANDIDATE_CELLS:
        val = _cell(rows, r, c)
        n = _to_float(val)
        n_int = int(n) if n > 0 else 0
        if n_int >= 100000 and str(n_int) != inn:
            tg_id = n_int
            break

    meta = {
        'inn': inn,
        'tg_id': tg_id,
        'name': str(_cell(rows, *META_FIELDS['name']) or '').strip(),
        'region': str(_cell(rows, *META_FIELDS['region']) or '').strip(),
        'district': str(_cell(rows, *META_FIELDS['district']) or '').strip(),
        'category': str(_cell(rows, *META_FIELDS['category']) or '').strip(),
        'manager': str(_cell(rows, *META_FIELDS['manager']) or '').strip(),
        'manager_phone': str(_cell(rows, *META_FIELDS['manager_phone']) or '').strip(),
        'manager_username': str(_cell(rows, *META_FIELDS['manager_username']) or '').strip().lstrip('@'),
    }

    # Парсим строки проектов до ОБЩЕЕ
    projects = []
    totals_row = None
    r = PROJECTS_START_ROW
    while r <= len(rows):
        num_cell = _cell(rows, r, COL_NUMBER)
        name_cell = _cell(rows, r, COL_NAME)

        # Конец секции проектов: ОБЩЕЕ
        if (isinstance(num_cell, str) and 'ОБЩЕЕ' in num_cell.upper()) or \
           (isinstance(name_cell, str) and 'ОБЩЕЕ' in name_cell.upper()):
            totals_row = r
            break

        # Пустая строка — считаем что проекты закончились
        if not name_cell or str(name_cell).strip() == '':
            r += 1
            continue

        # Собираем проект
        proj_months = {}
        for i, month_key in enumerate(MONTHS_RU):
            plan_col, fact_col, pct_col = COL_MONTHS[i]
            plan = _to_float(_cell(rows, r, plan_col))
            fact = _to_float(_cell(rows, r, fact_col))
            pct = _to_int_pct(_cell(rows, r, pct_col))
            proj_months[month_key] = {
                'plan': _fmt_money(plan),
                'fact': _fmt_money(fact),
                'plan_raw': plan,
                'fact_raw': fact,
                'percent': pct,
            }

        quarter_percent = _to_int_pct(_cell(rows, r, COL_QUARTER_PERCENT))
        quarter_plan = _to_float(_cell(rows, r, COL_QUARTER_PLAN))
        bonus_pct = _to_int_pct(_cell(rows, r, COL_BONUS_PCT))
        bonus_amount = _to_float(_cell(rows, r, COL_BONUS_AMOUNT))

        projects.append({
            'number': int(_to_float(num_cell)) if num_cell else None,
            'name': str(name_cell).strip(),
            'quarter_plan': _fmt_money(quarter_plan),
            'quarter_plan_raw': quarter_plan,
            'months': proj_months,
            'condition': str(_cell(rows, r, COL_CONDITION) or '').strip(),
            'percent': quarter_percent,
            'status': _status_for(quarter_percent),
            'remaining': str(_cell(rows, r, COL_REMAINING) or '').strip(),
            'bonus_percent': bonus_pct,
            'bonus_amount': _fmt_money(bonus_amount),
            'bonus_amount_raw': bonus_amount,
            # для совместимости со старым фронтом
            'fact': _fmt_money(sum(m['fact_raw'] for m in proj_months.values())),
            'plan': _fmt_money(quarter_plan),
        })
        r += 1

    # Итоги (ОБЩЕЕ)
    totals = {}
    if totals_row:
        total_quarter_plan = _to_float(_cell(rows, totals_row, COL_QUARTER_PLAN))
        total_months = {}
        for i, m in enumerate(MONTHS_RU):
            plan_col, fact_col, pct_col = COL_MONTHS[i]
            t_plan = _to_float(_cell(rows, totals_row, plan_col))
            t_fact = _to_float(_cell(rows, totals_row, fact_col))
            t_pct = _to_int_pct(_cell(rows, totals_row, pct_col))
            total_months[m] = {
                'plan': _fmt_money(t_plan),
                'fact': _fmt_money(t_fact),
                'percent': t_pct,
            }
        total_bonus = _to_float(_cell(rows, totals_row, COL_BONUS_AMOUNT))
        totals = {
            'quarter_plan': _fmt_money(total_quarter_plan),
            'quarter_plan_raw': total_quarter_plan,
            'months': total_months,
            'quarter_percent': _to_int_pct(_cell(rows, totals_row, COL_QUARTER_PERCENT)),
            'remaining': str(_cell(rows, totals_row, COL_REMAINING) or '').strip(),
            'total_bonus': _fmt_money(total_bonus),
            'total_bonus_raw': total_bonus,
        }

    # Агрегаты под существующий фронт
    stats = {'completed': 0, 'partial': 0, 'critical': 0}
    for p in projects:
        stats[p['status']] += 1

    bonus_earned = sum(p['bonus_amount_raw'] for p in projects)
    bonus_pending = sum(
        p['quarter_plan_raw'] - sum(m['fact_raw'] for m in p['months'].values())
        for p in projects
        if p['percent'] < 100
    )
    bonus_pending = max(0, bonus_pending)

    completed_names = ', '.join(p['name'] for p in projects if p['status'] == 'completed') or '—'

    return {
        # Богатые данные (новый формат)
        **meta,
        'city': meta['district'] or meta['region'],
        'projects': projects,
        'totals': totals,
        # Совместимость со старым фронтом
        'income_quarter': totals.get('total_bonus', '0'),
        'stats': stats,
        'months': totals.get('months', {}),
        'bonuses': {
            'accrued': {
                'amount': _fmt_money(bonus_earned),
                'desc': f"{stats['completed']} проектов на 100%+",
            },
            'potential': {
                'amount': '+ ' + _fmt_money(bonus_pending),
                'desc': f"До цели по {stats['partial'] + stats['critical']} проектам",
            },
            'completed': {
                'amount': _fmt_money(bonus_earned),
                'desc': completed_names,
            },
        },
    }


# ============================================================
# III-Q parser
# Широкая таблица: одна строка = одна аптека (юр.лицо, один ИНН).
# Колонки B..I — метаданные (ИНН, юр.название, сеть, кол-во точек, менеджер, категория).
# Колонки J.. — блоки по 10 колонок на проект:
#   проект | условия | план I-Q | план январь | факт январь
#         | план февраль | факт февраль | план март | факт март | ВП
# Блоки идут вправо до конца листа. Кол-во проектов — произвольное.
# ============================================================

IIIQ_SHEET_NAME = 'III-Q'
IIIQ_PROJECT_BLOCK_SIZE = 10

# Лист со связкой ИНН → telegram_id. Менеджер заполняет руками по мере выдачи
# доступа аптекам. 2 колонки: A=ИНН (число), B=telegram_id (число).
# Первая строка — заголовки, парсер её пропускает.
TG_IDS_SHEET_NAME = 'Доступы'

# Бонус % за проект (захардкожено по данным «Свод таб new», верхнего предела).
# Для проектов не из списка — фолбэк 7%.
# TODO: вынести в отдельный конфиг-лист в Google Sheets.
BONUS_PCT_BY_PROJECT = {
    'KRKA': 7, 'КРКА': 7,
    'KUSUM': 3, 'КУСУМ': 3,
    'WELFARM': 7, 'ВЕЛФАРМ': 7,
    'GETZ PHARMA': 7,
    'FERON': 7,
    'FARM STANDART': 7,
    'ZOMMER': 5,
    'ASFARMA': 7,
    'SYNERGY': 7,
    'COMPLETEPHARMA': 7,
    'BAYER': 7,
    'MEDEXPORT': 6,
    'ASTRA ZENECA': 6,
    'CCL': 7,
    'SAFE': 7,
    'SIRIUS': 7,
    'SENTISS': 7,
    'BIOMIND': 7,
}
DEFAULT_BONUS_PCT = 7


def _norm(s):
    return str(s or '').strip().lower()


def _parse_tg_id_mapping(rows):
    """
    Парсит лист 'Доступы'. Возвращает {inn: tg_id}.

    Структура:
      A1: ИНН | B1: telegram_id  (заголовки, пропускаются)
      A2+: ИНН | telegram_id     (значения)
    """
    result = {}
    if not rows or len(rows) < 2:
        return result
    for row in rows[1:]:
        if not row or len(row) < 2:
            continue
        inn_raw, tg_raw = row[0], row[1]
        if inn_raw in (None, '') or tg_raw in (None, ''):
            continue
        try:
            inn = str(int(_to_float(inn_raw)))
            tg = int(_to_float(tg_raw))
        except (ValueError, TypeError):
            continue
        if len(inn) < 7 or tg < 100000:
            continue
        result[inn] = tg
    return result


def _merge_tg_ids(pharmacies_by_inn, tg_map, source_label='DASH'):
    """Подставляет tg_id в распарсенные данные аптек по маппингу ИНН → tg_id."""
    if not tg_map:
        return
    bound = 0
    for inn, data in pharmacies_by_inn.items():
        if inn in tg_map:
            data['tg_id'] = tg_map[inn]
            bound += 1
    print(f"  ✓ [{source_label}] Привязок к Telegram: {bound}/{len(pharmacies_by_inn)}")


def _find_iiiq_header_row(rows):
    """Ищет строку-заголовок: содержит "ИНН" и "проект". Возвращает индекс (1-based) или None."""
    for i, row in enumerate(rows[:25]):
        if not row:
            continue
        has_inn = any(_norm(c) == 'инн' for c in row)
        has_project = any(_norm(c) == 'проект' for c in row)
        if has_inn and has_project:
            return i + 1  # 1-based
    return None


def _bonus_pct_for(project_name):
    key = (project_name or '').strip().upper()
    return BONUS_PCT_BY_PROJECT.get(key, DEFAULT_BONUS_PCT)


def _parse_iiiq_project_block(block, project_name=None):
    """
    Парсит один блок проекта (10 ячеек).
    Возвращает dict проекта в том же формате что _parse_pharmacy_sheet, или None если данных нет.
    """
    name = project_name or str(block[0] or '').strip()
    condition = str(block[1] or '').strip()
    q_plan = _to_float(block[2])
    jan_plan = _to_float(block[3]); jan_fact = _to_float(block[4])
    feb_plan = _to_float(block[5]); feb_fact = _to_float(block[6])
    mar_plan = _to_float(block[7]); mar_fact = _to_float(block[8])
    vp_raw = block[9]

    if not name:
        return None
    # Пропускаем проекты, где у этой аптеки нет ни плана, ни факта
    total_fact_check = jan_fact + feb_fact + mar_fact
    if q_plan == 0 and total_fact_check == 0:
        return None

    # Месячные проценты считаем сами
    def _pct(plan, fact):
        if plan <= 0:
            return 0
        return int(round(fact / plan * 100))

    months = {
        'january':  {'plan': _fmt_money(jan_plan), 'fact': _fmt_money(jan_fact),
                     'plan_raw': jan_plan, 'fact_raw': jan_fact, 'percent': _pct(jan_plan, jan_fact)},
        'february': {'plan': _fmt_money(feb_plan), 'fact': _fmt_money(feb_fact),
                     'plan_raw': feb_plan, 'fact_raw': feb_fact, 'percent': _pct(feb_plan, feb_fact)},
        'march':    {'plan': _fmt_money(mar_plan), 'fact': _fmt_money(mar_fact),
                     'plan_raw': mar_plan, 'fact_raw': mar_fact, 'percent': _pct(mar_plan, mar_fact)},
    }

    total_fact = jan_fact + feb_fact + mar_fact

    # Квартальный процент: если в ячейке S есть значение — используем его (формат: дробь 0.83 или целое 83);
    # если пусто — считаем сами.
    if vp_raw is None or vp_raw == '':
        q_percent = _pct(q_plan, total_fact)
    else:
        # Excel хранит проценты как доли (0.83). Но если кто-то вписал руками "83" — тоже понимаем.
        v = _to_float(vp_raw)
        q_percent = int(round(v * 100)) if abs(v) <= 5 else int(round(v))

    remaining = max(0.0, q_plan - total_fact)
    bonus_pct = _bonus_pct_for(name)
    # Бонус начисляется при выполнении плана. Простая модель: bonus = q_plan × bonus_pct% (если выполнили).
    # Подробнее логика DATFO может отличаться — пока такой MVP.
    bonus_amount = q_plan * bonus_pct / 100 if q_percent >= 100 else 0.0

    return {
        'number': None,
        'name': name,
        'quarter_plan': _fmt_money(q_plan),
        'quarter_plan_raw': q_plan,
        'months': months,
        'condition': condition,
        'percent': q_percent,
        'status': _status_for(q_percent),
        'remaining': '+ ' + _fmt_money(remaining) if q_percent >= 100 else _fmt_money(remaining),
        'bonus_percent': bonus_pct,
        'bonus_amount': _fmt_money(bonus_amount),
        'bonus_amount_raw': bonus_amount,
        'fact': _fmt_money(total_fact),
        'plan': _fmt_money(q_plan),
    }


def parse_iiiq_sheet(rows):
    """
    Парсит III-Q. Возвращает {inn: pharmacy_data} в формате, совместимом с
    _parse_pharmacy_sheet (тот же шейп dashboard_data, что читает Mini App).

    Особенности:
      - tg_id: не в III-Q. В выходе будет None — затем _apply_dashboard_updates
        попытается сохранить запись без привязки владельца (см. upsert_pharmacy_dashboard_only).
    """
    header_row = _find_iiiq_header_row(rows)
    if not header_row:
        return {}

    header = rows[header_row - 1]

    # Найти колонки метаданных по заголовкам
    col_idx = {}
    for j, c in enumerate(header):
        n = _norm(c)
        if n == 'инн' and 'inn' not in col_idx: col_idx['inn'] = j
        elif 'юридическое' in n: col_idx['legal_name'] = j
        elif n == 'аптеки': col_idx['pharm_name'] = j
        elif 'кол-во' in n: col_idx['count'] = j
        elif n in ('менежер', 'менеджер'): col_idx['manager'] = j
        elif 'категория' in n and 'datfo' in n: col_idx['category'] = j

    if 'inn' not in col_idx:
        return {}

    # Найти начала блоков проектов: все колонки с "проект" в заголовке
    project_starts = [j for j, c in enumerate(header) if _norm(c) == 'проект']

    results = {}
    skip_inn_placeholders = {'7777', '0'}  # строка TOTAL

    for r in range(header_row, len(rows)):  # header_row уже 1-based, что соответствует индексу следующей строки
        row = rows[r]
        if not row:
            continue

        inn_raw = row[col_idx['inn']] if col_idx['inn'] < len(row) else None
        if inn_raw is None or inn_raw == '':
            continue
        try:
            inn_num = int(_to_float(inn_raw))
        except (ValueError, TypeError):
            continue
        if inn_num < 1000000:  # ИНН минимум 7 цифр; 7777/0 — placeholder TOTAL
            continue
        inn = str(inn_num)
        if inn in skip_inn_placeholders:
            continue

        def _cellv(key):
            j = col_idx.get(key)
            if j is None or j >= len(row): return ''
            return str(row[j] or '').strip()

        legal_name = _cellv('legal_name')
        pharm_name = _cellv('pharm_name') or legal_name
        manager = _cellv('manager')
        category = _cellv('category')

        # Проекты
        projects = []
        for ps in project_starts:
            block = row[ps:ps + IIIQ_PROJECT_BLOCK_SIZE]
            if len(block) < IIIQ_PROJECT_BLOCK_SIZE:
                # padded короткий хвост
                block = list(block) + [None] * (IIIQ_PROJECT_BLOCK_SIZE - len(block))
            proj = _parse_iiiq_project_block(block)
            if proj:
                projects.append(proj)

        # Агрегаты как в _parse_pharmacy_sheet
        stats = {'completed': 0, 'partial': 0, 'critical': 0}
        for p in projects:
            stats[p['status']] += 1

        bonus_earned = sum(p['bonus_amount_raw'] for p in projects)
        bonus_pending = sum(
            (p['quarter_plan_raw'] - sum(m['fact_raw'] for m in p['months'].values())) * p['bonus_percent'] / 100
            for p in projects
            if p['percent'] < 100
        )
        bonus_pending = max(0, bonus_pending)

        # totals.months — суммируем по всем проектам
        total_months = {}
        for m in MONTHS_RU:
            tp = sum(p['months'][m]['plan_raw'] for p in projects)
            tf = sum(p['months'][m]['fact_raw'] for p in projects)
            total_months[m] = {
                'plan': _fmt_money(tp),
                'fact': _fmt_money(tf),
                'percent': int(round(tf / tp * 100)) if tp > 0 else 0,
            }

        total_quarter_plan = sum(p['quarter_plan_raw'] for p in projects)
        total_quarter_fact = sum(sum(m['fact_raw'] for m in p['months'].values()) for p in projects)
        quarter_percent = int(round(total_quarter_fact / total_quarter_plan * 100)) if total_quarter_plan > 0 else 0

        totals = {
            'quarter_plan': _fmt_money(total_quarter_plan),
            'quarter_plan_raw': total_quarter_plan,
            'months': total_months,
            'quarter_percent': quarter_percent,
            'remaining': '',
            'total_bonus': _fmt_money(bonus_earned),
            'total_bonus_raw': bonus_earned,
        }

        completed_names = ', '.join(p['name'] for p in projects if p['status'] == 'completed') or '—'

        results[inn] = {
            'inn': inn,
            'tg_id': None,  # в III-Q его нет — апсертим без привязки
            'name': pharm_name,
            'legal_name': legal_name,
            'region': '',
            'district': '',
            'category': category,
            'manager': manager,
            'manager_phone': '',
            'manager_username': '',
            'city': '',
            'projects': projects,
            'totals': totals,
            'income_quarter': totals.get('total_bonus', '0'),
            'stats': stats,
            'months': totals.get('months', {}),
            'bonuses': {
                'accrued': {
                    'amount': _fmt_money(bonus_earned),
                    'desc': f"{stats['completed']} проектов на 100%+",
                },
                'potential': {
                    'amount': '+ ' + _fmt_money(bonus_pending),
                    'desc': f"До цели по {stats['partial'] + stats['critical']} проектам",
                },
                'completed': {
                    'amount': _fmt_money(bonus_earned),
                    'desc': completed_names,
                },
            },
        }

    return results


def _read_all_sheets():
    """
    Читает аптеки из DASHBOARD_SHEET_ID.

    Приоритет:
      1) Если есть лист 'III-Q' — парсим только его (одна строка = одна аптека, все
         проекты в колонках). Возвращаем сразу.
      2) Иначе fallback: читаем все листы как «Свод таб new» (один лист = одна аптека).
    """
    wb = _open_workbook()

    # --- III-Q: широкая таблица со всеми аптеками ---
    iiiq_ws = next((ws for ws in wb.worksheets() if ws.title == IIIQ_SHEET_NAME), None)
    if iiiq_ws:
        try:
            # III-Q может быть очень широкой (18 проектов × 10 колонок + meta) и длинной
            raw = iiiq_ws.get('A1:HZ3000', value_render_option='UNFORMATTED_VALUE')
            results = parse_iiiq_sheet(raw)
            if results:
                print(f"  ✓ {IIIQ_SHEET_NAME!r}: {len(results)} аптек")
                # Подтягиваем привязки tg_id, если есть лист 'Доступы'
                tg_ws = next((ws for ws in wb.worksheets() if ws.title == TG_IDS_SHEET_NAME), None)
                if tg_ws:
                    try:
                        tg_rows = tg_ws.get('A1:B2000', value_render_option='UNFORMATTED_VALUE')
                        tg_map = _parse_tg_id_mapping(tg_rows)
                        _merge_tg_ids(results, tg_map, source_label='DASH')
                    except Exception as e:
                        print(f"⚠️ [DASH] лист {TG_IDS_SHEET_NAME!r}: {type(e).__name__}: {e}")
                return results
            print(f"⚠️ [DASH] лист {IIIQ_SHEET_NAME!r}: парсер не нашёл данных, пробуем старый формат")
        except Exception as e:
            print(f"⚠️ [DASH] лист {IIIQ_SHEET_NAME!r}: {type(e).__name__}: {e}; пробуем старый формат")

    # --- Fallback: парсим каждый лист как одну аптеку ---
    results = {}
    for ws in wb.worksheets():
        try:
            raw = ws.get('A1:U50', value_render_option='UNFORMATTED_VALUE')
        except Exception as e:
            print(f"⚠️ [DASH] лист {ws.title!r}: ошибка чтения {e}")
            continue

        data = _parse_pharmacy_sheet(raw)
        if data:
            results[data['inn']] = data
            print(f"  ✓ {ws.title!r}: ИНН={data['inn']} ({len(data['projects'])} проектов)")

    return results


async def _apply_dashboard_updates(by_inn, source_label='DASH'):
    """
    Общий upsert для всех аптек.

    - Аптеки с известным tg_id (из «Свод таб new») → upsert_pharmacy_full,
      одновременно проставляет владельца Telegram.
    - Аптеки без tg_id (из III-Q) → upsert_pharmacy_dashboard_only,
      сохраняем dashboard_data, привязку владельца не трогаем.
    """
    ok_with_tg = 0
    ok_no_tg = 0
    errors = []
    per_pharm = {}  # inn -> {'name', 'tg_id'}
    for inn, data in by_inn.items():
        legal_name = data.get('legal_name') or data.get('name') or inn
        display_name = data.get('name') or inn
        try:
            if data.get('tg_id'):
                await upsert_pharmacy_full(
                    inn=inn,
                    owner_tg_id=data['tg_id'],
                    business_name=legal_name,
                    pharmacy_name=display_name,
                    dashboard_data=data,
                )
                ok_with_tg += 1
                per_pharm[inn] = {'name': display_name, 'tg_id': data['tg_id']}
            else:
                await upsert_pharmacy_dashboard_only(
                    inn=inn,
                    business_name=legal_name,
                    pharmacy_name=display_name,
                    dashboard_data=data,
                )
                ok_no_tg += 1
                per_pharm[inn] = {'name': display_name, 'tg_id': None}
        except Exception as e:
            errors.append({'inn': inn, 'error': str(e)})
            print(f"❌ [{source_label}] inn={inn}: {e}")

    ok = ok_with_tg + ok_no_tg
    msg = f"✅ [{source_label}] Обновлено: {ok}/{len(by_inn)}"
    if ok_no_tg:
        msg += f"  (без TG_ID: {ok_no_tg})"
    if errors:
        msg += f"  ❌ ошибок: {len(errors)}"
    print(msg)
    return {
        'total_sheets': len(by_inn),
        'updated': ok,
        'updated_with_tg': ok_with_tg,
        'updated_no_tg': ok_no_tg,
        'skipped_no_tg': [],  # для совместимости с bot.py (раньше использовалось)
        'errors': errors,
        'per_pharm': per_pharm,
    }


async def sync_dashboard():
    if not DASHBOARD_SHEET_ID:
        print("⚠️ [DASH] DASHBOARD_SHEET_ID не задан в .env — синк пропущен")
        return

    print("🔄 [DASH] Чтение Google-таблицы...")
    loop = asyncio.get_running_loop()
    try:
        by_inn = await loop.run_in_executor(None, _read_all_sheets)
    except Exception as e:
        print(f"❌ [DASH] Ошибка чтения Google: {type(e).__name__}: {e}")
        return

    if not by_inn:
        print("⚠️ [DASH] Ни одного листа с ИНН не найдено.")
        return

    await _apply_dashboard_updates(by_inn, source_label='DASH')


# ============================================================
# Импорт из локального Excel-файла (.xlsx).
# Менеджер шлёт файл в бот → bot.py качает → вызывает эту функцию.
# Формат листов — тот же что в Google Sheets («Свод таб»).
# ============================================================
# Имя листа, который менеджер обновляет вручную (меняя ИНН в C4).
# Только он считается источником истины при загрузке xlsx.
PRIMARY_SHEET_NAME = 'Свод таб new'


def _read_sheet_rows(ws):
    """Читает первые 50×21 ячеек листа в 2D-список значений."""
    raw = []
    for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=21, values_only=True):
        raw.append(list(row))
    return raw


def _read_iiiq_rows_from_excel(ws):
    """III-Q может быть очень широкой (200+ колонок) и длинной — читаем больший диапазон."""
    raw = []
    for row in ws.iter_rows(min_row=1, max_row=3000, min_col=1, max_col=300, values_only=True):
        raw.append(list(row))
    return raw


def _read_all_sheets_from_excel(file_path):
    """
    Парсит .xlsx. Стратегия:
      1) Если есть лист 'III-Q' — берём его (одна строка = одна аптека, массовый импорт).
      2) Иначе если есть PRIMARY_SHEET_NAME ('Свод таб new') — берём только его (одна аптека).
      3) Иначе fallback: пробуем все листы (старое поведение).
    """
    from openpyxl import load_workbook

    # data_only=True — берём вычисленные значения формул, а не сами формулы
    wb = load_workbook(file_path, data_only=True, read_only=True)
    results = {}

    # --- 1. III-Q (массовый импорт) ---
    iiiq_ws = next((ws for ws in wb.worksheets if ws.title == IIIQ_SHEET_NAME), None)
    if iiiq_ws:
        try:
            raw = _read_iiiq_rows_from_excel(iiiq_ws)
            results = parse_iiiq_sheet(raw)
            if results:
                print(f"  ✓ {IIIQ_SHEET_NAME!r}: {len(results)} аптек")
                # Подтягиваем привязки tg_id, если есть лист 'Доступы'
                tg_ws = next((ws for ws in wb.worksheets if ws.title == TG_IDS_SHEET_NAME), None)
                if tg_ws:
                    try:
                        tg_rows = []
                        for row in tg_ws.iter_rows(min_row=1, max_row=2000, min_col=1, max_col=2, values_only=True):
                            tg_rows.append(list(row))
                        tg_map = _parse_tg_id_mapping(tg_rows)
                        _merge_tg_ids(results, tg_map, source_label='XLSX')
                    except Exception as e:
                        print(f"⚠️ [XLSX] лист {TG_IDS_SHEET_NAME!r}: {type(e).__name__}: {e}")
                wb.close()
                return results
            print(f"⚠️ [XLSX] лист {IIIQ_SHEET_NAME!r}: парсер не нашёл данных, пробуем старый формат")
        except Exception as e:
            print(f"⚠️ [XLSX] лист {IIIQ_SHEET_NAME!r}: {type(e).__name__}: {e}; пробуем старый формат")

    # --- 2. Свод таб new (одна аптека) → 3. все листы ---
    primary_ws = next((ws for ws in wb.worksheets if ws.title == PRIMARY_SHEET_NAME), None)
    targets = [primary_ws] if primary_ws else list(wb.worksheets)

    for ws in targets:
        try:
            raw = _read_sheet_rows(ws)
        except Exception as e:
            print(f"⚠️ [XLSX] лист {ws.title!r}: ошибка чтения {e}")
            continue

        data = _parse_pharmacy_sheet(raw)
        if data:
            results[data['inn']] = data
            print(f"  ✓ {ws.title!r}: ИНН={data['inn']} ({len(data['projects'])} проектов)")

    wb.close()
    return results


async def sync_dashboard_from_excel(file_path):
    """Импортирует все аптеки из локального .xlsx-файла. Возвращает summary."""
    print(f"🔄 [XLSX] Чтение {file_path}...")
    loop = asyncio.get_running_loop()
    try:
        by_inn = await loop.run_in_executor(None, _read_all_sheets_from_excel, file_path)
    except Exception as e:
        print(f"❌ [XLSX] Ошибка чтения файла: {type(e).__name__}: {e}")
        return {'error': f'{type(e).__name__}: {e}', 'updated': 0}

    if not by_inn:
        return {'error': 'Ни одного листа с ИНН не найдено', 'updated': 0}

    return await _apply_dashboard_updates(by_inn, source_label='XLSX')


if __name__ == "__main__":
    asyncio.run(sync_dashboard())
