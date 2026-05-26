"""
Локальный тест парсера на dattfo (3).xlsx — не лезет в Google.
Используем openpyxl, скармливаем те же rows в _parse_pharmacy_sheet.

    uv run python test_parse_excel.py
"""
import json
from openpyxl import load_workbook
from dashboard_sync import _parse_pharmacy_sheet

wb = load_workbook('dattfo (3).xlsx', data_only=True)

for name in wb.sheetnames:
    ws = wb[name]
    rows = []
    for r in range(1, min(ws.max_row + 1, 51)):
        row = []
        for c in range(1, min(ws.max_column + 1, 22)):
            row.append(ws.cell(row=r, column=c).value)
        rows.append(row)

    data = _parse_pharmacy_sheet(rows)
    if data:
        print(f"\n=== Лист {name!r} распарсен ===")
        print(json.dumps(data, ensure_ascii=False, indent=2, default=str))
    else:
        print(f"--- лист {name!r}: пропущен (нет ИНН в C4) ---")
