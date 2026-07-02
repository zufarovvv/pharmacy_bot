"""
Импорт истории выплат (карточка счёта 5110 «Оплата за услуги трейд-маркетинга»).

Файл — статичная выгрузка из 1С за год, формат листа TDSheet:
  Период | Документ | Аналитика Дт | Сумма
«Аналитика Дт» многострочная: первая строка — название фирмы ("ALIF PHARMA" MCHJ),
дальше договор. ИНН в файле НЕТ — матчим к pharmacies по названию фирмы
(business_name / legal_name / name), нормализуя кавычки, орг-формы и пробелы.

Запуск:
  uv run python payments_import.py "/путь/к/оплата_за_услуги_трейд_маркетинга_2025г.xlsx"

Пишет в таблицу payments (полная перезагрузка: TRUNCATE + INSERT).
"""
import asyncio
import re
import sys
from collections import defaultdict
from datetime import datetime

from dotenv import load_dotenv
from openpyxl import load_workbook

from database import get_connection, replace_all_payments

load_dotenv()

SHEET_NAME = 'TDSheet'
DATE_RE = re.compile(r'^\d{2}\.\d{2}\.\d{4}$')
# Орг-формы (латиница и кириллица), которые выкидываем при нормализации имени
ORG_RE = re.compile(
    r"\b(MCHJ|MChJ|XK|XX|OK|QK|OOO|ООО|ЧП|СП|ХК|МЧЖ|ЧК|КК|ФХ|ОК)\b\.?",
    re.IGNORECASE)


def norm_name(s):
    """Нормализует название фирмы: первая строка, без кавычек/орг-форм/лишних пробелов."""
    if not s:
        return ''
    s = str(s).split('\n')[0]
    s = s.upper().replace('’', "'").replace('`', "'").replace('«', '"').replace('»', '"')
    s = re.sub(r'["“”]', '', s)
    s = ORG_RE.sub(' ', s)
    s = re.sub(r"[^A-ZА-Я0-9' -]", ' ', s)
    return re.sub(r'[\s-]+', ' ', s).strip()


def parse_payments_xlsx(path):
    """Читает карточку 5110. Возвращает список платежей."""
    ws = load_workbook(path, data_only=True)[SHEET_NAME]
    pays = []
    for r in range(1, ws.max_row + 1):
        d, doc, ana, amt = (ws.cell(r, c).value for c in range(1, 5))
        if not (isinstance(d, str) and DATE_RE.match(d.strip())):
            continue
        if amt is None or ana is None:
            continue
        try:
            amount = float(str(amt).replace(' ', '').replace(',', '.'))
        except ValueError:
            continue
        pays.append({
            'date': datetime.strptime(d.strip(), '%d.%m.%Y').date(),
            'company': str(ana).split('\n')[0].strip(),
            'key': norm_name(ana),
            'doc': str(doc or '').split('\n')[0].strip()[:200],
            'amount': amount,
        })
    return pays


async def build_name_index():
    """Индексы норм.имя → {inn}: точный и «слитный» (без пробелов) как фолбэк."""
    conn = await get_connection()
    try:
        rows = await conn.fetch(
            "SELECT inn, business_name, dashboard_data->>'legal_name' AS legal, "
            "dashboard_data->>'name' AS pname FROM pharmacies")
    finally:
        await conn.close()

    exact, packed = defaultdict(set), defaultdict(set)
    for r in rows:
        for cand in (r['business_name'], r['legal'], r['pname']):
            k = norm_name(cand)
            if k:
                exact[k].add(r['inn'])
                packed[k.replace(' ', '')].add(r['inn'])
    return exact, packed


def match_inn(key, exact, packed):
    """ИНН по нормализованному имени; None если нет или неоднозначно (>1 ИНН)."""
    inns = exact.get(key) or packed.get(key.replace(' ', ''))
    if inns and len(inns) == 1:
        return next(iter(inns))
    return None


async def import_payments(path):
    pays = parse_payments_xlsx(path)
    if not pays:
        print('⚠️ [PAY] В файле не найдено ни одного платежа — формат не распознан')
        return {'total': 0, 'matched': 0}

    exact, packed = await build_name_index()

    records, unmatched = [], defaultdict(lambda: [0, 0.0])
    ambiguous = defaultdict(lambda: [0, 0.0])
    for p in pays:
        inns = exact.get(p['key']) or packed.get(p['key'].replace(' ', ''))
        if inns and len(inns) == 1:
            records.append((next(iter(inns)), p['company'], p['date'], p['amount'], p['doc']))
        elif inns:
            ambiguous[p['company']][0] += 1
            ambiguous[p['company']][1] += p['amount']
        else:
            unmatched[p['company']][0] += 1
            unmatched[p['company']][1] += p['amount']

    await replace_all_payments(records)

    total_sum = sum(p['amount'] for p in pays)
    matched_sum = sum(r[3] for r in records)
    print(f"  ✓ [PAY] Импортировано платежей: {len(records)}/{len(pays)} "
          f"({matched_sum:,.0f} из {total_sum:,.0f} сум, {matched_sum / total_sum * 100:.1f}%)")
    if ambiguous:
        print(f"  ⚠️ [PAY] Неоднозначные имена (одно имя — несколько ИНН), пропущены: {len(ambiguous)}")
        for name, (cnt, s) in sorted(ambiguous.items(), key=lambda x: -x[1][1])[:5]:
            print(f"      • {name}: {cnt} плат., {s:,.0f}")
    if unmatched:
        print(f"  ⚠️ [PAY] Не найдены в БД: {len(unmatched)} компаний "
              f"({sum(v[0] for v in unmatched.values())} плат., "
              f"{sum(v[1] for v in unmatched.values()):,.0f} сум) — топ-10:")
        for name, (cnt, s) in sorted(unmatched.items(), key=lambda x: -x[1][1])[:10]:
            print(f"      • {name}: {cnt} плат., {s:,.0f}")
    return {'total': len(pays), 'matched': len(records)}


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    asyncio.run(import_payments(sys.argv[1]))
