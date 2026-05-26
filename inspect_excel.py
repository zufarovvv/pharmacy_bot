"""
Инспектор структуры Excel. Запусти и пришли вывод:

    uv run python inspect_excel.py
"""
from openpyxl import load_workbook

FILE = 'dattfo (3).xlsx'

wb = load_workbook(FILE, data_only=True)  # data_only=True -> читаем вычисленные значения, не формулы

print(f"=== {FILE} ===")
print(f"Всего листов: {len(wb.sheetnames)}\n")

for i, name in enumerate(wb.sheetnames):
    ws = wb[name]
    print(f"[{i}] sheet={name!r}  rows={ws.max_row}  cols={ws.max_column}")

print("\n" + "=" * 60)
print("ДЕТАЛЬНЫЙ ДАМП КАЖДОГО ЛИСТА (первые 35 строк, 18 колонок)")
print("=" * 60)

for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n----- SHEET: {name!r} -----")
    for row_idx in range(1, min(ws.max_row + 1, 36)):
        cells = []
        for col_idx in range(1, min(ws.max_column + 1, 19)):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                cells.append("·")
            else:
                s = str(v).strip()
                if len(s) > 22:
                    s = s[:20] + "…"
                cells.append(s)
        # покажем только если в строке есть хоть что-то
        if any(c != "·" for c in cells):
            print(f"  R{row_idx:2}: " + " | ".join(cells))
